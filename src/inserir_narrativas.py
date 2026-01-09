from openpyxl import load_workbook
import re

def inserir_narrativa(
    caminho_planilha_modelo: str,
    caminho_saida: str,
) -> None:
    """
    A partir da terceira linha, verifica o tamanho da coluna SAP123
    e, se for maior que 144 caracteres, escreve "see basic data text"
    na coluna Narrativa.

    :param caminho_planilha_modelo: Caminho da planilha de entrada
    :param caminho_saida: Caminho onde a planilha será salva
    """
    print("Atualizando Narrativa por tamanho de SAP123...")

    wb = load_workbook(caminho_planilha_modelo)
    ws = wb.active

    col_sap123 = None
    col_narrativa = None

    for coluna in ws.iter_cols(min_row=1, max_row=1):
        for celula in coluna:
            valor = celula.value
            if valor is None:
                continue
            nome = re.sub(r"\s+", "", str(valor)).upper()
            if nome == "SAP123":
                col_sap123 = celula.column
            elif nome == "NARRATIVA":
                col_narrativa = celula.column

    if col_sap123 is None or col_narrativa is None:
        print("Aviso: colunas SAP123 ou Narrativa não encontradas no cabeçalho.")
        return

    alteradas = 0
    exemplos = []
    for linha in range(3, ws.max_row + 1):
        valor = ws.cell(row=linha, column=col_sap123).value
        if isinstance(valor, str) and len(valor) > 141:
            ws.cell(row=linha, column=col_narrativa).value = "verificar internal comment"
            alteradas += 1
            if len(exemplos) < 5:
                exemplos.append(linha)

    wb.save(caminho_saida)
    print(f"Narrativa atualizada por tamanho: {alteradas} linhas")


