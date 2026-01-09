from openpyxl import load_workbook

def inserir_valores_fixos(
    caminho_planilha_modelo: str,
    caminho_saida: str,
) -> None:
    """
    Insere valores fixos nas colunas SAP10 e SAP14
    - SAP10: valor "10" em linhas com código na primeira coluna
    - SAP14: valor "NDB" em linhas com código na primeira coluna
    
    :param caminho_planilha_modelo: Caminho da planilha de entrada
    :param caminho_saida: Caminho onde a planilha será salva
    """
    print("Inserindo valores fixos em SAP10/SAP1...")

    # Carrega a planilha
    workbook = load_workbook(caminho_planilha_modelo)
    planilha = workbook.active
    
    # Encontra o índice das colunas SAP10 e SAP14
    col_sap10 = None
    col_sap1 = None
    
    # Percorre o cabeçalho (primeira linha) para encontrar as colunas
    for coluna in planilha.iter_cols(min_row=1, max_row=1):
        for celula in coluna:
            if celula.value == 'SAP10':
                col_sap10 = celula.column
            elif celula.value == 'SAP14':
                col_sap1 = celula.column
    
    # Se as colunas não existem, aborta com aviso
    if col_sap10 is None or col_sap1 is None:
        print("Aviso: colunas SAP10 ou SAP1 não encontradas no cabeçalho.")
        return
    
    alteradas = 0
    exemplos = []
    # Percorre as linhas a partir da terceira linha (pulando cabeçalhos)
    for linha_num in range(3, planilha.max_row + 1):
        # Verifica se há código na primeira coluna (coluna A)
        codigo = planilha[f'A{linha_num}'].value
        
        if codigo is not None and codigo != '':  # Se tem código
            # Insere "10" na coluna SAP10
            planilha.cell(row=linha_num, column=col_sap10).value = '10'
            # Insere "NDB" na coluna SAP14
            planilha.cell(row=linha_num, column=col_sap1).value = 'NDB'
            alteradas += 1
            if len(exemplos) < 5:
                exemplos.append(linha_num)
    
    # Salva a planilha
    workbook.save(caminho_saida)
    print(f"Valores fixos aplicados: {alteradas} linhas")