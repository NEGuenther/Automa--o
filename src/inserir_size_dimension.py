from openpyxl import load_workbook

def inserir_size_dimension(
    caminho_planilha_modelo: str,
    caminho_saida: str,
) -> None:
    """
    A partir da terceira linha, verifica o tamanho da coluna SAP123
    e, se for maior que 144 caracteres, escreve "see basic data text"
    na coluna SAP15.

    :param caminho_planilha_modelo: Caminho da planilha de entrada
    :param caminho_saida: Caminho onde a planilha será salva
    """
    wb = load_workbook(caminho_planilha_modelo)
    ws = wb.active

    col_sap123 = None
    col_sap15 = None

    for coluna in ws.iter_cols(min_row=1, max_row=1):
        for celula in coluna:
            valor = celula.value
            if valor is None:
                continue
            nome = str(valor).replace(" ", "").upper()
            if nome == "SAP123":
                col_sap123 = celula.column
            elif nome == "SAP15":
                col_sap15 = celula.column

    if col_sap123 is None or col_sap15 is None:
        print("Colunas SAP123 ou SAP15 não encontradas na planilha")
        return

    for linha in range(3, ws.max_row + 1):
        valor = ws.cell(row=linha, column=col_sap123).value
        if isinstance(valor, str) and len(valor) > 144:
            ws.cell(row=linha, column=col_sap15).value = "see basic data text"

    wb.save(caminho_saida)
    print(f"Atualização de SAP15 concluída em {caminho_saida}")
